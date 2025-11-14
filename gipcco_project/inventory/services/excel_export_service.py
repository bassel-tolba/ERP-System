# gipcco_project/inventory/services/excel_export_service.py
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import date
from decimal import Decimal


def generate_excel_response(filename: str, title: str, headers: list, data: list, totals: list = None, merge_title: bool = True):
    """
    Generic function to generate an Excel file from report data.
    """
    wb = Workbook()
    ws = wb.active
    # Use a safe sheet title
    ws.title = title[:31].replace('/', '-').replace('\\', '-') # Excel sheet title limit is 31 chars, no slashes
    ws.sheet_view.rightToLeft = True

    # --- Styles ---
    title_font = Font(name='Calibri', size=16, bold=True)
    header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(horizontal='right', vertical='center')
    text_align = Alignment(horizontal='left', vertical='center')
    total_font = Font(name='Calibri', size=12, bold=True)
    total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    banded_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Main Title ---
    if merge_title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        header_row_index = 3
    else:
        header_row_index = 1

    # --- Header ---
    ws.insert_rows(header_row_index -1, amount=1) if header_row_index > 1 else None
    for i, header_text in enumerate(headers, 1):
        cell = ws.cell(row=header_row_index, column=i, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[header_row_index].height = 25

    # Freeze header row
    ws.freeze_panes = ws.cell(row=header_row_index + 1, column=1)

    # --- Data Rows ---
    start_row = header_row_index + 1
    for row_idx, row_data in enumerate(data):
        current_row = start_row + row_idx
        for col_idx, cell_data in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=cell_data)
            cell.border = thin_border
            
            # Formatting based on data type
            if isinstance(cell_data, (int, float, Decimal)):
                cell.alignment = cell_align
                cell.number_format = '#,##0.00' if isinstance(cell_data, (float, Decimal)) else '#,##0'
            else:
                cell.alignment = text_align

            # Apply banding
            if row_idx % 2 != 0:
                cell.fill = banded_fill

    # --- Totals Row ---
    if totals:
        total_row_index = ws.max_row + 1
        for i, total_data in enumerate(totals, 1):
            cell = ws.cell(row=total_row_index, column=i, value=total_data)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border
            if isinstance(total_data, (int, float, Decimal)):
                cell.alignment = cell_align
                cell.number_format = '#,##0.00' if isinstance(total_data, (float, Decimal)) else '#,##0'
            else:
                cell.alignment = text_align if i == 1 else cell_align

    # --- Auto-fit columns ---
    column_widths = {}
    for i, header in enumerate(headers, 1):
        column_widths[i] = len(header) + 2 # Start with header length

    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            if cell.value:
                cell_len = len(str(cell.value))
                if isinstance(cell.value, (int, float, Decimal)):
                     cell_len += (cell_len // 3) # Account for thousand separators
                column_widths[cell.column] = max(column_widths.get(cell.column, 0), cell_len + 4)

    for i, width in column_widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(width, 50) # Cap width at 50

    # --- Create HTTP Response ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response

def export_ar_aging_to_excel(report_data, totals, as_of_date):
    filename = f"AR_Aging_Report_{as_of_date}"
    title = "AR Aging Report"
    headers = [
        "Customer", "Current", "1-30 Days", "31-60 Days", "61-90 Days",
        "91-120 Days", "Over 120 Days", "Total Due", "Unapplied Credits", "Net Balance"
    ]
    
    data = []
    for item in report_data:
        data.append([
            item['customer'].name,
            item['current'],
            item['1_30'],
            item['31_60'],
            item['61_90'],
            item['91_120'],
            item['over_120'],
            item['total_due'],
            item['unapplied_credits'],
            item['net_balance']
        ])
        
    totals_list = [
        "Total",
        totals['current'],
        totals['1_30'],
        totals['31_60'],
        totals['61_90'],
        totals['91_120'],
        totals['over_120'],
        totals['total_due'],
        totals['unapplied_credits'],
        totals['net_balance']
    ]
    return generate_excel_response(filename, f"AR Aging Report as of {as_of_date}", headers, data, totals_list)

def export_customer_statement_to_excel(customer, start_date, end_date, opening_balance, transactions, closing_balance):
    filename = f"Customer_Statement_{customer.name}_{start_date}_to_{end_date}"
    title = "Customer Statement"
    headers = ["Date", "Transaction", "Details", "Debit", "Credit", "Balance"]
    
    data = [
        ["Opening Balance", "", "", "", "", opening_balance]
    ]
    for t in transactions:
        data.append([
            t['date'],
            "Invoice" if t['type'] == 'invoice' else "Payment",
            f"#{t['obj'].invoice_number}" if t['type'] == 'invoice' else t['obj'].description,
            t['debit'],
            t['credit'],
            t['balance']
        ])
    
    totals = ["Closing Balance", "", "", "", "", closing_balance]
    return generate_excel_response(filename, f"Statement for {customer.name}", headers, data, totals)


def export_sales_by_customer_to_excel(report_data, grand_total, start_date, end_date):
    filename = f"Sales_by_Customer_{start_date}_to_{end_date}"
    title = "Sales by Customer Report"
    headers = ["Customer Name", "Invoice Count", "Total Sales"]
    
    data = [
        [item['customer__name'], item['invoice_count'], item['total_sales']]
        for item in report_data
    ]
    
    totals_list = ["Grand Total", "", grand_total]
    return generate_excel_response(filename, f"Sales by Customer ({start_date} to {end_date})", headers, data, totals_list)


def export_sales_by_product_to_excel(report_data, grand_total, start_date, end_date):
    filename = f"Sales_by_Product_{start_date}_to_{end_date}"
    title = "Sales by Product Report"
    headers = ["Product Name", "Unit", "Total Quantity Sold", "Total Revenue"]
    
    data = [
        [item['product_name'], item['unit'], item['total_quantity_sold'], item['total_revenue']]
        for item in report_data
    ]
    
    totals_list = ["Grand Total", "", "", grand_total]
    return generate_excel_response(filename, f"Sales by Product ({start_date} to {end_date})", headers, data, totals_list)


def export_sales_order_backlog_to_excel(report_data):
    filename = f"Sales_Order_Backlog_{date.today()}"
    title = "Sales Order Backlog Report"
    headers = [
        "SO Number", "Order Date", "Customer Name", "Product",
        "Quantity Ordered", "Quantity Shipped", "Quantity Backlog", "Unit"
    ]
    
    data = [
        [
            item['so_number'],
            item['order_date'],
            item['customer_name'],
            item['product_name'],
            item['quantity_ordered'],
            item['quantity_shipped'],
            item['quantity_backlog'],
            item['unit']
        ]
        for item in report_data
    ]
    return generate_excel_response(filename, f"Sales Order Backlog as of {date.today()}", headers, data)
